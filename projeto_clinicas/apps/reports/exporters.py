"""
Exportadores de relatorio: CSV, Excel e PDF.

CSV funciona sem dependencia externa. Excel usa ``openpyxl`` e PDF usa
``reportlab`` -- ambos opcionais: se a biblioteca nao estiver instalada, a
exportacao correspondente informa o motivo em vez de quebrar a aplicacao.
"""
from __future__ import annotations

import csv
import io
from typing import Iterable, List, Sequence

from django.http import HttpResponse


def export_csv(filename: str, headers: Sequence[str], rows: Iterable[Sequence]) -> HttpResponse:
    response = HttpResponse(content_type="text/csv; charset=utf-8-sig")
    response["Content-Disposition"] = f'attachment; filename="{filename}.csv"'
    response.write("﻿")  # BOM para o Excel reconhecer acentuacao
    writer = csv.writer(response, delimiter=";")
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    return response


def export_xlsx(
    filename: str, headers: Sequence[str], rows: Iterable[Sequence], title: str = "Relatorio"
) -> HttpResponse:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return _missing_dependency("openpyxl", "Excel")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title[:31]
    sheet.append(list(headers))
    header_fill = PatternFill("solid", start_color="0B5ED7")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    for row in rows:
        sheet.append(list(row))
    for index, header in enumerate(headers, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=index).column_letter].width = max(
            14, min(48, len(str(header)) + 6)
        )
    buffer = io.BytesIO()
    workbook.save(buffer)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
    return response


def _clinic_chrome(clinic):
    """
    Monta o cabecalho (logo/nome/documento/endereco/telefone da clinica) e o
    rodape ("<Plataforma> - Relatorio gerado em ... - Pagina X de Y") usados
    em todo PDF de relatorio quando ``clinic`` e informado.

    Retorna ``(header_fn, canvas_class)`` prontos para
    ``SimpleDocTemplate.build(..., onFirstPage=header_fn,
    onLaterPages=header_fn, canvasmaker=canvas_class)``. A contagem total de
    paginas so e conhecida depois que todo o documento foi montado, entao o
    rodape usa o padrao classico do reportlab de canvas que guarda o estado
    de cada pagina e so desenha o rodape no final (``save``).
    """
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.utils import ImageReader
    from reportlab.pdfgen import canvas as canvas_module
    from django.conf import settings as dj_settings
    from django.utils import timezone

    platform_name = getattr(dj_settings, "PLATFORM_NAME", "Alume Tech")

    logo_reader = None
    if clinic.logo:
        try:
            clinic.logo.open("rb")
            logo_reader = ImageReader(io.BytesIO(clinic.logo.read()))
        except Exception:
            logo_reader = None
        finally:
            try:
                clinic.logo.close()
            except Exception:
                pass

    contact_parts = [part for part in [clinic.document, clinic.full_address] if part]
    if clinic.phone:
        contact_parts.append(f"Tel: {clinic.phone}")
    contact_line = "  -  ".join(contact_parts)

    def header_fn(c, doc):
        width, height = doc.pagesize
        text_x = 12 * mm
        c.saveState()
        if logo_reader is not None:
            try:
                c.drawImage(
                    logo_reader, 12 * mm, height - 22 * mm, width=16 * mm, height=16 * mm,
                    preserveAspectRatio=True, mask="auto",
                )
                text_x = 32 * mm
            except Exception:
                pass
        c.setFont("Helvetica-Bold", 11)
        c.drawString(text_x, height - 12 * mm, str(clinic.trade_name or clinic.legal_name))
        if contact_line:
            c.setFont("Helvetica", 7)
            c.drawString(text_x, height - 17 * mm, contact_line)
        c.setStrokeColor(colors.HexColor("#cccccc"))
        c.line(12 * mm, height - 24 * mm, width - 12 * mm, height - 24 * mm)
        c.restoreState()

    class ClinicPageCanvas(canvas_module.Canvas):
        def __init__(self, *args, **kwargs):
            canvas_module.Canvas.__init__(self, *args, **kwargs)
            self._jja_page_states = []

        def showPage(self):
            self._jja_page_states.append(dict(self.__dict__))
            self._startPage()

        def save(self):
            total_pages = len(self._jja_page_states)
            for state in self._jja_page_states:
                self.__dict__.update(state)
                self._jja_draw_footer(total_pages)
                canvas_module.Canvas.showPage(self)
            canvas_module.Canvas.save(self)

        def _jja_draw_footer(self, total_pages):
            width, _height = self._pagesize
            generated = timezone.localtime(timezone.now()).strftime("%d/%m/%Y %H:%M")
            footer = (
                f"{platform_name} - Relatorio gerado em {generated} - "
                f"Pagina {self._pageNumber} de {total_pages}"
            )
            self.saveState()
            self.setFont("Helvetica", 7)
            self.setFillColor(colors.HexColor("#666666"))
            self.drawCentredString(width / 2, 8 * mm, footer)
            self.restoreState()

    return header_fn, ClinicPageCanvas


def export_pdf(
    filename: str,
    headers: Sequence[str],
    rows: Iterable[Sequence],
    title: str = "Relatorio",
    subtitle: str = "",
    clinic=None,
) -> HttpResponse:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )
    except ImportError:
        return _missing_dependency("reportlab", "PDF")

    build_kwargs = {}
    top_margin = 12 * mm
    if clinic is not None:
        header_fn, canvas_class = _clinic_chrome(clinic)
        build_kwargs = {
            "onFirstPage": header_fn,
            "onLaterPages": header_fn,
            "canvasmaker": canvas_class,
        }
        top_margin = 28 * mm

    buffer = io.BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=top_margin,
        bottomMargin=16 * mm,
        title=title,
    )
    styles = getSampleStyleSheet()
    elements: List = [Paragraph(title, styles["Title"])]
    if subtitle:
        elements.append(Paragraph(subtitle, styles["Normal"]))
    elements.append(Spacer(1, 8))

    data = [list(headers)] + [[str(value) for value in row] for row in rows]
    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0b5ed7")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cccccc")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f4f6fb")]),
            ]
        )
    )
    elements.append(table)
    document.build(elements, **build_kwargs)

    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}.pdf"'
    return response


def export_record_pdf(
    filename: str,
    patient,
    entry_rows: Sequence[Sequence],
    prescription_rows: Sequence[Sequence],
    exam_rows: Sequence[Sequence],
    clinic=None,
) -> HttpResponse:
    """
    PDF do prontuario completo de um paciente: atendimentos, prescricoes e
    exames em secoes separadas (ao contrario de ``export_pdf``, que exporta
    uma unica tabela). Reaproveita a mesma marca da clinica (cabecalho/
    rodape) via ``_clinic_chrome``.
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError:
        return _missing_dependency("reportlab", "PDF")

    build_kwargs = {}
    top_margin = 14 * mm
    if clinic is not None:
        header_fn, canvas_class = _clinic_chrome(clinic)
        build_kwargs = {
            "onFirstPage": header_fn,
            "onLaterPages": header_fn,
            "canvasmaker": canvas_class,
        }
        top_margin = 28 * mm

    buffer = io.BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        topMargin=top_margin,
        bottomMargin=16 * mm,
        title=f"Prontuario - {patient}",
    )
    styles = getSampleStyleSheet()
    elements: List = [
        Paragraph(f"Prontuario - {patient}", styles["Title"]),
        Spacer(1, 10),
    ]

    def section(section_title, headers, rows):
        elements.append(Paragraph(section_title, styles["Heading3"]))
        if not rows:
            elements.append(Paragraph("Nenhum registro.", styles["Normal"]))
            elements.append(Spacer(1, 8))
            return
        data = [list(headers)] + [[str(value) for value in row] for row in rows]
        table = Table(data, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0b5ed7")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cccccc")),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f4f6fb")]),
                ]
            )
        )
        elements.append(table)
        elements.append(Spacer(1, 10))

    section("Atendimentos", ["Data", "Profissional", "Modelo", "Situacao"], entry_rows)
    section("Prescricoes", ["Data", "Profissional", "Tipo", "Itens"], prescription_rows)
    section("Exames", ["Solicitado em", "Profissional", "Situacao"], exam_rows)

    document.build(elements, **build_kwargs)

    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}.pdf"'
    return response


def _missing_dependency(package: str, label: str) -> HttpResponse:
    return HttpResponse(
        f"Exportacao em {label} indisponivel: a biblioteca '{package}' nao esta "
        f"instalada neste ambiente. Instale com 'pip install {package}'.",
        status=501,
        content_type="text/plain; charset=utf-8",
    )


def export(
    fmt: str, filename: str, headers, rows, title: str = "Relatorio", subtitle: str = "",
    clinic=None,
):
    fmt = (fmt or "csv").lower()
    if fmt == "xlsx":
        return export_xlsx(filename, headers, rows, title)
    if fmt == "pdf":
        return export_pdf(filename, headers, rows, title, subtitle, clinic=clinic)
    return export_csv(filename, headers, rows)
